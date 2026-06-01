"""
generar_catalogo_maestro.py
Genera catalogo_articulos_por_area.js y .json
a partir de los Excel de inventario Enero-Abril 2026.
Solo extrae: codigo, nombre, area. Sin precios ni existencias.

REGLAS (acordadas con el usuario):
1. Mismo codigo con nombres distintos entre meses -> nombre canonico = ABRIL 2026.
2. Mismo nombre con codigos distintos -> NO se resuelve, solo se reporta.
3. Articulo con nombre pero SIN codigo -> NO entra al catalogo, solo al reporte.
4. Codigo fuera del prefijo esperado del area -> se conserva tal cual, pero se reporta.

La tabla de cada Excel termina en la fila de autosuma (TOTAL); todo lo que
sigue (firmas: ELABORO, C.P., JEFE DE DEPARTAMENTO...) se ignora.
"""

import json
import re
from pathlib import Path
from collections import defaultdict
import openpyxl

REPO_DIR = Path(r"C:\Users\sauli\OneDrive\Documentos\REPOSITORIO\files")
DOWNLOADS = Path(r"C:\Users\sauli\Downloads")
ENERO_DIR = Path(r"C:\Users\sauli\OneDrive\Documentos\claude ai\INV. ENERO CIERRE")
OUTPUT_DIR = REPO_DIR / "data" / "catalogos"

# Tokens que NO son ni codigo ni nombre de articulo
KEYWORDS_SKIP = {
    "CODIGO", "CÓDIGO", "COD", "COD.", "DESCRIPCION", "DESCRIPCIÓN",
    "ARTICULO", "ARTÍCULO", "NOMBRE", "TOTAL", "TOTALES", "SUBTOTAL",
    "SUB TOTAL", "NONE", "", "N/A", "NA",
}

# Marcadores que indican fin de la tabla (autosuma) o pie de firma
END_MARKERS = (
    "TOTAL", "TOTALES", "GRAN TOTAL", "TOTAL GENERAL", "SUMA",
    "ELABOR", "REVIS", "AUTORIZ", "VO. BO", "V° B°", "VO.BO",
    "NOMBRE Y FIRMA", "JEFE DE DEPARTAMEN", "JEFA DE DEPARTAMEN",
)

# Prefijo de codigo esperado por area
PREFIJO_AREA = {
    "Snack": "S-",
    "Tienda": "T-",
    "Farmacia": "F-",
    "Nutrición": "2220",
    "Mantenimiento": "M-",
}

# Orden cronologico: el ultimo (Abril) gana como nombre canonico.
FILES_POR_AREA = {
    "Snack": [
        ENERO_DIR / "01.ENERO.2026 INVENTARIO.FINAL.SNACK...xlsx",
        DOWNLOADS / "02.FEBRERO.2026 INVENTARIO DE SNACK (1).xlsx",
        DOWNLOADS / "03.MARZO.2026 INVENTARIO DE SNACK (3).xlsx",
        REPO_DIR / "inventarios" / "04.ABRIL.2026.INVENTARIO.SNACK.FINAL.xlsx",
    ],
    "Tienda": [
        ENERO_DIR / "01.ENERO.2026 INVENTARIO.FINAL.TIENDA RECUERDOS..xlsx",
        DOWNLOADS / "02.FEBRERO.2026 INVENTARIO DE TIENDA (1).xlsx",
        DOWNLOADS / "03.MARZO.2026.INVENTARIO.TIENDARECUERDOS.FINAL. (2).xlsx",
        REPO_DIR / "inventarios" / "04.ABRIL.2026.INVENTARIO.TIENDARECUERDOS.FINAL.xlsx",
    ],
    "Farmacia": [
        ENERO_DIR / "01.ENERO.2026 INVENTARIO.FINAL.FARMACIA..xlsx",
        DOWNLOADS / "02.FEBRERO.2026 INVENTARIO DE FARMACIA.xlsx",
        DOWNLOADS / "03.MARZO.2026.INVENTARIO.FARMACIA.FINAL.17-04-2026.xlsx",
        REPO_DIR / "inventarios" / "04.ABRIL.2026.INVENTARIO.FARMACIA.FINAL.xlsx",
    ],
    "Nutrición": [
        ENERO_DIR / "01.ENERO.2026 INVENTARIO.FINAL.NUTRICION...xlsx",
        DOWNLOADS / "02.FEBRERO.2026 INVENTARIO NUTRICION (2).xlsx",
        DOWNLOADS / "03.MARZO.2026.INVENTARIO.NUTRICION.FINAL.xlsx",
        REPO_DIR / "inventarios" / "04.ABRIL.2026.INVENTARIO.NUTRICION.FINAL+CIE (2).xlsx",
    ],
    "Mantenimiento": [
        ENERO_DIR / "01.ENERO.2026 INVENTARIO.FINAL.MANTENIMIENTO.xlsx",
        DOWNLOADS / "02.FEBRERO.2026 INVENTARIO MANTENIMIENTO.JULIA (3).xlsx",
        DOWNLOADS / "03.MARZO.2026 INVENTARIO MANTENIMIENTO FINAL ZOOLOGICO(CRISTY) (2).xlsx",
        REPO_DIR / "inventarios" / "04.ABRIL.2026.MANTENIMIENTO.FINAL.xlsx",
    ],
}

MESES = ["Enero", "Febrero", "Marzo", "Abril"]


def normalizar(texto):
    if texto is None:
        return ""
    return re.sub(r"\s+", " ", str(texto)).strip()


def find_best_sheet(wb):
    """Prioriza hoja con 'INV' y 'FINAL'; si no, la activa."""
    for name in wb.sheetnames:
        if name.strip().upper() == "INV. FINAL":
            return wb[name]
    for name in wb.sheetnames:
        n = name.upper()
        if "INV" in n and "FINAL" in n:
            return wb[name]
    return wb.active


def find_header(ws):
    """Detecta fila de encabezado y columnas de codigo y nombre."""
    for r in range(1, 30):
        row = [normalizar(ws.cell(r, c).value).upper() for c in range(1, 25)]
        for ci, val in enumerate(row, 1):
            if val in ("CODIGO", "CÓDIGO"):
                nom_col = None
                for ci2, val2 in enumerate(row, 1):
                    if val2 in ("DESCRIPCION", "DESCRIPCIÓN", "ARTICULO",
                                "ARTÍCULO", "NOMBRE"):
                        nom_col = ci2
                        break
                if nom_col:
                    return r, ci, nom_col
    return None, None, None


def es_fin_de_tabla(ws, r, cc, nc):
    """True si la fila r marca el fin de la tabla (autosuma) o pie de firma."""
    # Revisa las primeras columnas hasta nombre+1 por algun marcador de cierre
    for c in range(1, nc + 2):
        val = normalizar(ws.cell(r, c).value).upper()
        if not val:
            continue
        for marker in END_MARKERS:
            if val.startswith(marker):
                return True
    return False


def extraer_articulos(path):
    """
    Devuelve (validos, sin_codigo):
      validos    = lista de (codigo, nombre) con codigo y nombre presentes
      sin_codigo = lista de nombres con nombre presente pero codigo ausente
    Lee solo el cuerpo de la tabla (corta en la fila de autosuma/firmas).
    """
    validos, sin_codigo = [], []
    try:
        # NO read_only: acceso aleatorio con ws.cell() es O(n^2) en ese modo.
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        ws = find_best_sheet(wb)
        hr, cc, nc = find_header(ws)
        if hr is None:
            print(f"  [WARN] Sin encabezado detectado: {path.name}", flush=True)
            wb.close()
            return validos, sin_codigo

        blancos_seguidos = 0
        for r in range(hr + 1, ws.max_row + 1):
            if es_fin_de_tabla(ws, r, cc, nc):
                break  # autosuma o firmas -> fin de tabla
            cod = normalizar(ws.cell(r, cc).value)
            nom = normalizar(ws.cell(r, nc).value)
            tiene_cod = bool(cod) and cod.upper() not in KEYWORDS_SKIP
            tiene_nom = bool(nom) and nom.upper() not in KEYWORDS_SKIP

            if not tiene_cod and not tiene_nom:
                blancos_seguidos += 1
                # 25 filas vacias seguidas -> fin de tabla (respaldo)
                if blancos_seguidos >= 25:
                    break
                continue
            blancos_seguidos = 0

            if tiene_cod and tiene_nom:
                validos.append((cod, nom))
            elif tiene_nom and not tiene_cod:
                sin_codigo.append(nom)  # regla 3: nombre sin codigo -> solo reporte
        wb.close()
    except Exception as e:
        print(f"  [ERROR] {path.name}: {e}", flush=True)
    return validos, sin_codigo


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    catalogo = {}
    reporte = {}
    fuentes_log = []

    for area, archivos in FILES_POR_AREA.items():
        print(f"\n=== {area} ===", flush=True)
        # Para nombre canonico ABRIL: guardamos el nombre por mes mas reciente.
        nombre_por_codigo = {}        # codigo -> nombre (gana el mes mas reciente)
        pares_todos = []              # {codigo, nombre} con duplicados (reporte)
        sin_codigo_todos = []         # nombres sin codigo (reporte)

        for idx, path in enumerate(archivos):
            mes = MESES[idx] if idx < len(MESES) else f"archivo{idx}"
            if not path.exists():
                print(f"  [SKIP] No existe ({mes}): {path.name}", flush=True)
                continue
            validos, sin_cod = extraer_articulos(path)
            print(f"  [{mes}] {path.name}: {len(validos)} con codigo, "
                  f"{len(sin_cod)} sin codigo", flush=True)
            fuentes_log.append({
                "area": area, "mes": mes, "archivo": path.name,
                "con_codigo": len(validos), "sin_codigo": len(sin_cod),
            })
            for cod, nom in validos:
                pares_todos.append({"codigo": cod, "nombre": nom})
                # Orden cronologico Ene->Abr: sobrescribir hace que ABRIL gane.
                nombre_por_codigo[cod] = nom
            sin_codigo_todos.extend(sin_cod)

        # --- Catalogo final: un objeto por codigo, nombre canonico (Abril) ---
        lista_final = [{"codigo": cod, "nombre": nombre_por_codigo[cod]}
                       for cod in sorted(nombre_por_codigo)]
        catalogo[area] = lista_final

        # --- Reporte de inconsistencias ---
        por_codigo = defaultdict(set)
        por_nombre = defaultdict(set)
        for par in pares_todos:
            por_codigo[par["codigo"]].add(par["nombre"])
            por_nombre[par["nombre"]].add(par["codigo"])

        conflicto_nombre_por_codigo = {
            cod: {"nombres_vistos": sorted(noms),
                  "nombre_canonico_abril": nombre_por_codigo.get(cod)}
            for cod, noms in por_codigo.items() if len(noms) > 1
        }
        conflicto_codigo_por_nombre = {
            nom: sorted(cods) for nom, cods in por_nombre.items() if len(cods) > 1
        }
        prefijo = PREFIJO_AREA[area]
        fuera_prefijo = sorted(
            {cod for cod in nombre_por_codigo if not cod.startswith(prefijo)}
        )
        # nombres sin codigo, unicos y ordenados
        sin_codigo_unicos = sorted(set(sin_codigo_todos))

        reporte[area] = {
            "prefijo_esperado": prefijo,
            "articulos_unicos": len(lista_final),
            "conflicto_nombre_por_codigo": {
                "total": len(conflicto_nombre_por_codigo),
                "detalle": conflicto_nombre_por_codigo,
            },
            "conflicto_codigo_por_nombre": {
                "total": len(conflicto_codigo_por_nombre),
                "detalle": conflicto_codigo_por_nombre,
            },
            "sin_codigo": {
                "total": len(sin_codigo_unicos),
                "detalle": sin_codigo_unicos,
            },
            "fuera_de_prefijo": {
                "total": len(fuera_prefijo),
                "detalle": [{"codigo": c, "nombre": nombre_por_codigo[c]}
                            for c in fuera_prefijo],
            },
        }
        print(f"  -> {len(lista_final)} unicos | "
              f"conf.nombre/codigo={len(conflicto_nombre_por_codigo)} | "
              f"conf.codigo/nombre={len(conflicto_codigo_por_nombre)} | "
              f"sin_codigo={len(sin_codigo_unicos)} | "
              f"fuera_prefijo={len(fuera_prefijo)}", flush=True)

    # --- Escribir salidas ---
    json_path = OUTPUT_DIR / "catalogo_articulos_por_area.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False, indent=2)

    js_path = OUTPUT_DIR / "catalogo_articulos_por_area.js"
    with open(js_path, "w", encoding="utf-8") as f:
        f.write("// Catalogo maestro Enero-Abril 2026 - solo codigo y nombre por area\n")
        f.write("// Nombre canonico = Abril 2026. Generado automaticamente - no editar a mano.\n")
        f.write("window.CATALOGO_ARTICULOS_POR_AREA = ")
        json.dump(catalogo, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    reporte_path = OUTPUT_DIR / "reporte_inconsistencias.json"
    with open(reporte_path, "w", encoding="utf-8") as f:
        json.dump({"fuentes": fuentes_log, "areas": reporte}, f,
                  ensure_ascii=False, indent=2)

    # --- Resumen final en consola ---
    print("\n\n===== RESUMEN FINAL =====", flush=True)
    tot = dict(unicos=0, cnc=0, ccn=0, sc=0, fp=0)
    for area, r in reporte.items():
        tot["unicos"] += r["articulos_unicos"]
        tot["cnc"] += r["conflicto_nombre_por_codigo"]["total"]
        tot["ccn"] += r["conflicto_codigo_por_nombre"]["total"]
        tot["sc"] += r["sin_codigo"]["total"]
        tot["fp"] += r["fuera_de_prefijo"]["total"]
        print(f"{area:14} unicos={r['articulos_unicos']:5} | "
              f"conf.nom/cod={r['conflicto_nombre_por_codigo']['total']:3} | "
              f"conf.cod/nom={r['conflicto_codigo_por_nombre']['total']:3} | "
              f"sin_codigo={r['sin_codigo']['total']:3} | "
              f"fuera_pref={r['fuera_de_prefijo']['total']:2}", flush=True)
    print(f"{'TOTAL':14} unicos={tot['unicos']:5} | "
          f"conf.nom/cod={tot['cnc']:3} | conf.cod/nom={tot['ccn']:3} | "
          f"sin_codigo={tot['sc']:3} | fuera_pref={tot['fp']:2}", flush=True)
    print(f"\n[OK] {json_path}")
    print(f"[OK] {js_path}")
    print(f"[OK] {reporte_path}")


if __name__ == "__main__":
    main()
